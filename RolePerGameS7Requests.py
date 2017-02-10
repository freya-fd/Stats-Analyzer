import os
import time
import threading
import requests
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import cell
from functools import wraps

ENDPOINT = 'https://na.api.pvp.net/api/lol/na/'
KEY = os.environ["DEV_KEY"]
API_KEY = 'api_key=' + KEY
MATCHES = []
BEGIN_TIME = '?beginTime=1481108400000'
#pylint: disable=E0001, E1101, C0111, C0103

def rate_limited(max_per_10sec):
    lock = threading.Lock()
    min_interval = 10.0 / max_per_10sec

    def decorate(func):
        last_time_called = time.perf_counter()

        @wraps(func)
        def rate_limited_function(*args, **kwargs):
            lock.acquire()
            nonlocal last_time_called
            elapsed = time.perf_counter() - last_time_called
            left_to_wait = min_interval - elapsed

            if left_to_wait > 0:
                time.sleep(left_to_wait)

            ret = func(*args, **kwargs)
            last_time_called = time.perf_counter()
            lock.release()
            return ret

        return rate_limited_function

    return decorate

def get_champ_name(id):
    r = requests.get('https://global.api.pvp.net/api/lol/static-data/na/v1.2/champion/' + str(id) + '?' + API_KEY)
    return r.json()['name']

def id_from_name(name):
    r = requests.get(ENDPOINT + 'v1.4/summoner/by-name/' + name + '?' + API_KEY)
    return r.json()[name]['id']

def get_matchlist(mId, season, queue):
    # if season == 7:
    #     s = 'SEASON2016'
    # if queue == 'solo':
    #     q = 'TEAM_BUILDER_RANKED_SOLO'
    # elif queue == 'flex':
    #     q = 'RANKED_FLEX_SR'
    # r = requests.get(ENDPOINT + 'v2.2/matchlist/by-summoner/' + mId + '?rankedQueues='
    #                 + QUEUE + '&seasons=' + SEASON + API_KEY)
    r = requests.get(ENDPOINT + 'v2.2/matchlist/by-summoner/' + mId + BEGIN_TIME + '&' + API_KEY)
    return r.json()['matches']
    # for x in MATCHLIST:
    #     MATCHES.append(x['matchId'])

@rate_limited(10)
def get_match_info(mId, sn):
    r = requests.get(ENDPOINT + 'v2.2/match/' + mId + '?' + API_KEY)
    duration = r.json()['matchDuration']
    participantIds = r.json()['participantIdentities']
    participants = r.json()['participants']
    for p in participantIds:
        if p['player']['summonerName'] == sn:
            pId = p['participantId']
    for p in participants:
        if p['participantId'] == pId:
            if duration <= 300:
                return 'Remake'
            elif p['stats']['winner']:
                return 'Win'
            else:
                return 'Loss'

def main():
    # summonerName = input('Enter the summoner you would like to get data for: ')
    # print('Looking up ID ', summonerName)
    SUMID = str(id_from_name('brimstoner'))
    # print('ID for ' + summonerName + ' is ' + SUMID)
    # print('\nGetting matchlist for ' + summonerName)
    # seasonInput = input('\nEnter the season to get data for (as a number): ')
    # queueInput = input('\nEnter the queue type to get data for (flex or solo): ')
    # print('\nGetting season ' + seasonInput + ' ' + queueInput + ' queue data for ' + summonerName)
    # get_matchlist(SUMID, seasonInput, queueInput)
    MATCHLIST = get_matchlist(SUMID, 7, 'solo')

    wb = Workbook()
    dest_filename = 'Brimstoner' + '_role_per_game.xlsx'

    ws1 = wb.active
    ws1.title = "Roles Per Game"

    ws1.cell(column=1, row=1, value="MatchID")
    ws1.cell(column=2, row=1, value="Champion")
    ws1.cell(column=3, row=1, value="Role")
    ws1.cell(column=4, row=1, value="Lane")
    ws1.cell(column=5, row=1, value="Queue")
    ws1.cell(column=6, row=1, value="W/L")

    for row, match in enumerate(MATCHLIST, start=2):
        ws1.cell(column=1, row=row, value=match['matchId'])
        ws1.cell(column=2, row=row, value=get_champ_name(match['champion']))
        ws1.cell(column=3, row=row, value=match['role'])
        ws1.cell(column=4, row=row, value=match['lane'])
        ws1.cell(column=5, row=row, value=match['queue'])
        ws1.cell(column=6, row=row, value=get_match_info(str(match['matchId']), 'Brimstoner'))

    wb.save(filename=dest_filename)

if __name__ == "__main__":
    main()
