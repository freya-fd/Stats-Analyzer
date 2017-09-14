import os
import time
import threading
import requests
import json
import glob
import config as cfg
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import cell
from openpyxl.styles import NamedStyle
from functools import wraps
from pprint import pprint
from enum import Enum

ENDPOINT = 'https://na1.api.riotgames.com/lol/' # might need na. for older summoners
KEY = cfg.DEV_KEY
API_KEY = 'api_key=' + KEY
CHAMP_LIST = {}
DIR = os.path.dirname(__file__)
CURR_HEADER = {}
#pylint: disable=E0001, E1101, C0111, C0103

#TODO: Refactor load_champ_data and surrounding functions to remove from main
#TODO: Develop algorithm to parse role/lane and ensure (as well as possible) accurateness. If has smite, is jg for example. Check champion tags, items maybe?
#TODO: Add input to choose Solo, Flex or All

def parse_header(header):
    method_limit = header['X-Method-Rate-Limit']
    method_count = header['X-Method-Rate-Limit-Count']
    method_limit_10 = method_limit.split(':')[0]
    method_count_10 = method_count.split(':')[0]
    if 'X-App-Rate-Limit' in header:
        app_limits = header['X-App-Rate-Limit']
        app_counts = header['X-App-Rate-Limit-Count']        
        app_limit_120 = app_limits.split(',')[0].split(':')[0]
        app_limit_1 = app_limits.split(',')[1].split(':')[0]
        app_count_120 = app_counts.split(',')[0].split(':')[0]
        app_count_1 = app_counts.split(',')[1].split(':')[0]        
        return {
            'app_limit_120':int(app_limit_120),
            'app_limit_1':int(app_limit_1),
            'app_count_120':int(app_count_120),
            'app_count_1':int(app_count_1),
            'method_limit_10':int(method_limit_10),
            'method_count_10':int(method_count_10)
            }
    else:
        return {
            'app_limit_120':100,
            'app_limit_1':20,
            'app_count_120':0,
            'app_count_1':0,
            'method_limit_10':int(method_limit_10),
            'method_count_10':int(method_count_10)
            }

def rate_limiter(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        global CURR_HEADER
        if CURR_HEADER == {}:
            ret = func(*args, **kwargs)
        else:
            limit_data = parse_header(CURR_HEADER)
            if (limit_data['app_limit_1'] - limit_data['app_count_1']) < 2:
                time.sleep(1)
                ret = func(*args, **kwargs)
            elif (limit_data['method_limit_10'] - limit_data['method_count_10']) < 2:
                time.sleep(10)
                ret = func(*args, **kwargs)
            elif (limit_data['app_limit_120'] - limit_data['app_count_120']) < 2:
                time.sleep(80)
                ret = func(*args, **kwargs)
            else:
                ret = func(*args, **kwargs)
        return ret
    return wrapper

@rate_limiter
def get_request(url):
    global CURR_HEADER
    r = requests.get(url)
    CURR_HEADER = r.headers
    print(CURR_HEADER)
    return r.json()

class Queue(Enum):
    ARAM = '65'
    INVASION = '990'
    FLEX = '440'
    SOLO = '420'

def load_champ_data():
    print("load_champ_data")
    with open(os.path.join(DIR, 'Data\\champ_list.txt')) as f:
        global CHAMP_LIST
        CHAMP_LIST = json.load(f)

# Returns true if champ_list.txt is current
def check_version():
    print("check_version")
    r = get_request(ENDPOINT + 'static-data/v3/versions?' + API_KEY)    
    return(r[0] == CHAMP_LIST['version'])

def get_champ_list():
    print("get_champ_list")
    if check_version() == True:
        return(CHAMP_LIST['data'])
    else:
        r = get_request(ENDPOINT + 'static-data/v3/champions?locale=en_US&dataById=true&' + API_KEY)
        with open(os.path.join(DIR, 'Data\\champ_list.txt'), 'w') as f:
            json.dump(r, f)
        return(CHAMP_LIST['data'])

def get_champ_name(id, champ_list):
    print("get_champ_name")
    return(champ_list[str(id)]['name'])

def duplicate_match(summoner_name, queue, old_data, match_id):
    print('duplicate_match')
    if old_data == []:
        return False
    else: # if a file for this summoner/queue exists
        for m in old_data:
            if m['gameId'] == match_id:
                return True
        return False

def get_current_match_data(summoner_name, queue):
    print('get_current_match_data')
    if glob.glob(os.path.join(DIR, 'Data\\' + summoner_name + '_' + queue + '_' + 'match_list.txt')) == []:
        return []
    else:
        with open(os.path.join(DIR, 'Data\\' + summoner_name + '_' + queue + '_' + 'match_list.txt')) as f:
            return json.load(f)

def update_match_data(data, summoner_name, queue):
    print("update_match_data")
    with open(os.path.join(DIR, 'Data\\' + summoner_name + '_' + queue + '_' + 'match_list.txt'), 'w') as f:
        json.dump(data, f)

def id_from_name(name):
    print("id_from_name")
    r = get_request(ENDPOINT + 'summoner/v3/summoners/by-name/' + name + '?' + API_KEY)
    return r['accountId']

def get_matchlist(accountID, season, queue):
    print("get_matchlist")
    begin_index = 0
    accountID = str(accountID)
    r = get_request(ENDPOINT + 'match/v3/matchlists/by-account/' + accountID + '?season=' + season + '&queue=' + queue + '&beginIndex=' + str(begin_index) + '&' + API_KEY)
    matches = r['matches']
    total_games = r['totalGames']
    remaining_games = total_games
    while total_games > begin_index:
        begin_index += 100
        r = get_request(ENDPOINT + 'match/v3/matchlists/by-account/' + accountID + '?season=' + season + '&queue=' + queue + '&beginIndex=' + str(begin_index) + '&' + API_KEY)
        total_games = r['totalGames']
        matches.extend(r['matches'])
    return matches

def get_match_info(game_id):
    print('get_match_info')
    game_id = str(game_id)
    r = get_request(ENDPOINT + 'match/v3/matches/' + game_id + '?' + API_KEY)
    return r

def get_team(match, current_account_id):
    print('get_team')
    participant_identities = match['participantIdentities']
    for x in participant_identities:
        if x['player']['currentAccountId'] == current_account_id:
            if x['participantId'] <= 5:
                return 'blue'
            elif x['participantId'] > 5:
                return 'red'
            
def get_teammates(match, current_account_id, team):
    print('get_teammates')
    participant_identities = match['participantIdentities']
    teammates = []
    for x in participant_identities:
        if team == 'blue':
            if x['player']['currentAccountId'] != current_account_id and x['participantId'] <= 5:
                teammates.append(x['player']['summonerName'])
        elif team == 'red':
            if x['player']['currentAccountId'] != current_account_id and x['participantId'] > 5:
                teammates.append(x['player']['summonerName'])
    return teammates

def get_match_state(match, team):
    print('get_match_state')
    if match['gameDuration'] < 300:
        return 'Remake'
    if team == 'blue':
        state = match['teams'][0]['win']
    elif team == 'red':
        state = match['teams'][1]['win']
    if state == 'Fail':
        return 'Loss'
    elif state == 'Win':
        return 'Win'

def make_workbook(summoner_name, matchlist, accountID, queue, old_data):
    print('make_workbook')
    global CHAMP_LIST
    dest_filename = os.path.join(DIR, summoner_name.lower() + '_' + queue.lower() + '_role_per_game.xlsx')
    if old_data == []:
        date_style = NamedStyle(name='datetime', number_format='M/D/YYYY HH:MM AM/PM')
        wb = Workbook()        
        ws1 = wb.active
        ws1.title = 'Roles Per Game - ' + queue
        ws1.cell(column=1, row=1, value='GameID')
        ws1.cell(column=2, row=1, value='Date')
        ws1.cell(column=3, row=1, value='Champion')
        ws1.cell(column=4, row=1, value='Role')
        ws1.cell(column=5, row=1, value='Lane')
        ws1.cell(column=6, row=1, value='Queue')
        ws1.cell(column=7, row=1, value='W/L')
        ws1.cell(column=8, row=1, value='Teammate1')
        ws1.cell(column=9, row=1, value='Teammate2')
        ws1.cell(column=10, row=1, value='Teammate3')
        ws1.cell(column=11, row=1, value='Teammate4')
    else:
        wb = load_workbook(dest_filename)
        ws1 = wb.active
        date_style = ws1.cell(column=2, row=2).style

    for row, match in enumerate(reversed(matchlist), start=2):
            if not duplicate_match(summoner_name, queue, old_data, match['gameId']):
                match_info = get_match_info(match['gameId'])
                team = get_team(match_info, accountID)
                teammates = get_teammates(match_info, accountID, team)
                ws1.cell(column=1, row=row, value=match['gameId'])
                excel_time = (((match['timestamp'] / 1000) - 18000) / 86400) + 25568 # Converts unix-time to excel date/time
                ws1.cell(column=2, row=row, value=excel_time)
                ws1.cell(column=2, row=row).style = date_style
                ws1.cell(column=3, row=row, value=get_champ_name(match['champion'], CHAMP_LIST))
                ws1.cell(column=4, row=row, value=match['role'])
                ws1.cell(column=5, row=row, value=match['lane'])
                ws1.cell(column=6, row=row, value='Solo')
                ws1.cell(column=7, row=row, value=get_match_state(match_info, team))
                ws1.cell(column=8, row=row, value=teammates[0])
                ws1.cell(column=9, row=row, value=teammates[1])
                ws1.cell(column=10, row=row, value=teammates[2])
                ws1.cell(column=11, row=row, value=teammates[3])

    wb.save(filename=dest_filename)

def main():
    global CHAMP_LIST
    load_champ_data()
    summoner_name = input('Enter the summoner you would like to get data for: ')
    accountID = id_from_name(summoner_name)
    print('Getting matchlist for ' + summoner_name)

    # seasonInput = input('\nEnter the season to get data for (as a number): ')
    # queueInput = input('\nEnter the queue type to get data for (flex or solo): ')
    # print('\nGetting season ' + seasonInput + ' ' + queueInput + ' queue data for ' + summoner_name)

    matchlist = get_matchlist(accountID, '9', Queue.SOLO.value)
    old_data = get_current_match_data(summoner_name, Queue.SOLO.name)
    update_match_data(matchlist, summoner_name, Queue.SOLO.name)
    CHAMP_LIST = get_champ_list()

    make_workbook(summoner_name, matchlist, accountID, Queue.SOLO.name, old_data)

if __name__ == "__main__":
    main()
