import os
import time
import threading
import requests
import json
import config as cfg
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import cell
from functools import wraps
from pprint import pprint
from enum import Enum

ENDPOINT = 'https://na1.api.riotgames.com/lol/' # might need na. for older summoners
KEY = cfg.DEV_KEY
API_KEY = 'api_key=' + KEY
MATCHES = []
CHAMP_LIST = {}
myTeammates = []
matchPlayers = []
DIR = os.path.dirname(__file__)
#pylint: disable=E0001, E1101, C0111, C0103

#TODO: Include functionality to check the match_list file for existing data and only pull new data

class Queue(Enum):
    ARAM = '65'
    INVASION = '990'
    FLEX = '440'
    SOLO = '420'

def rate_limited(max_per_second):
    lock = threading.Lock()
    min_interval = 1.0 / max_per_second

    def decorate(func):
        last_time_called = time.perf_counter()

        @wraps(func)
        def rate_limited_function(*args, **kwargs):
            lock.acquire()
            nonlocal last_time_called
            elapsed = time.perf_counter() - last_time_called
            left_to_wait = min_interval - elapsed

            if left_to_wait > 0:
                time.sleep(left_to_wait)

            ret = func(*args, **kwargs)
            last_time_called = time.perf_counter()
            lock.release()
            return ret

        return rate_limited_function

    return decorate

def load_champ_data():
     with open(os.path.join(DIR, 'Data\\champ_list.txt')) as data_file:
        global CHAMP_LIST
        CHAMP_LIST = json.load(data_file)

# Returns true if champ_list.txt is current
def check_version():
    r = requests.get(ENDPOINT + 'static-data/v3/versions?' + API_KEY)
    print(r.json())
    return(r.json()[0] == CHAMP_LIST['version'])

def get_champ_list():
    if check_version() == True:
        return(CHAMP_LIST['data'])
    else:
        r = requests.get(ENDPOINT + 'static-data/v3/champions?locale=en_US&dataById=true&' + API_KEY)
        print(r.headers)
        with open(os.path.join(DIR, 'Data\\champ_list.txt'), 'w') as f:
            json.dump(r.json(), f)
        return(CHAMP_LIST['data'])

def get_champ_name(id):
    return(get_champ_list()[str(id)]['name'])

def update_match_data(data):
    with open(os.path.join(DIR, 'Data\\match_list.txt'), 'w') as f:
        json.dump(data, f)

def id_from_name(name):
    r = requests.get(ENDPOINT + 'summoner/v3/summoners/by-name/' + name + '?' + API_KEY)
    return str(r.json()['accountId'])

@rate_limited(1)
def get_matchlist(accountID, season, queue):
    begin_index = 0
    r = requests.get(ENDPOINT + 'match/v3/matchlists/by-account/' + accountID + '?season=' + season + '&queue=' + queue + '&beginIndex=' + str(begin_index) + '&' + API_KEY)
    matches = r.json()['matches']
    total_games = r.json()['totalGames']
    remaining_games = total_games
    while remaining_games > 0:
        begin_index += 100
        remaining_games -= 100
        r = requests.get(ENDPOINT + 'match/v3/matchlists/by-account/' + accountID + '?season=' + season + '&queue=' + queue + '&beginIndex=' + str(begin_index) + '&' + API_KEY)
        matches.extend(r.json()['matches'])
    return matches

# -----------------------------------------^^NEW AND TESTED^^-----------------------------------------------------------

def get_teammates(pdict, team, myid):
    global matchPlayers
    matchPlayers = []
    for playerid, sumName in pdict.items():
        if team == 'blue' and myid != playerid and playerid <= 5:
            myTeammates.append(sumName)
            matchPlayers.append(sumName)
        elif team == 'red' and myid != playerid and playerid >= 6:
            myTeammates.append(sumName)
            matchPlayers.append(sumName)

@rate_limited(1)
def get_match_info(mId, sn):
    r = requests.get(ENDPOINT + 'v2.2/match/' + mId + '?' + API_KEY)
    print('Current count: ' + r.headers['X-Method-Rate-Limit-Count'] + 'out of ' + r.headers['X-Method-Rate-Limit'])
    duration = r.json()['matchDuration']
    participantIds = r.json()['participantIdentities']
    participants = r.json()['participants']
    pdict = {}
    for p in participantIds:
        pdict[p['participantId']] = p['player']['summonerName']
        if p['player']['summonerName'] == sn:
            pId = p['participantId']
            if pId <=5:
                team = 'blue'
            else:
                team = 'red'
    get_teammates(pdict, team, pId)
    for p in participants:
        if p['participantId'] == pId:
            if duration <= 300:
                return 'Remake'
            elif p['stats']['winner']:
                return 'Win'
            else:
                return 'Loss'

def main():
    load_champ_data()
    summonerName = input('Enter the summoner you would like to get data for: ')
    # print('Looking up ID for ' + summonerName)
    # accountID = id_from_name(summonerName)
    # print('ID for ' + summonerName + ' is ' + accountID)
    # print('Getting matchlist for ' + summonerName)

    # seasonInput = input('\nEnter the season to get data for (as a number): ')
    # queueInput = input('\nEnter the queue type to get data for (flex or solo): ')
    # print('\nGetting season ' + seasonInput + ' ' + queueInput + ' queue data for ' + summonerName)

    # matchlist = get_matchlist(accountID, '9', Queue.SOLO.value)
    matchlist = get_matchlist('460742', '9', Queue.SOLO.value)
    update_match_data(matchlist)

    wb = Workbook()
    dest_filename = summonerName.lower() + '_role_per_game.xlsx'

    ws1 = wb.active
    ws1.title = 'Roles Per Game'

    ws1.cell(column=1, row=1, value='GameID')
    ws1.cell(column=2, row=1, value='Date')
    ws1.cell(column=3, row=1, value='Champion')
    ws1.cell(column=4, row=1, value='Role')
    ws1.cell(column=5, row=1, value='Lane')
    ws1.cell(column=6, row=1, value='Queue')
    ws1.cell(column=7, row=1, value='W/L')
    ws1.cell(column=8, row=1, value='Teammate1')
    ws1.cell(column=9, row=1, value='Teammate2')
    ws1.cell(column=10, row=1, value='Teammate3')
    ws1.cell(column=11, row=1, value='Teammate4')

    for row, match in enumerate(reversed(matchlist), start=2):
        ws1.cell(column=1, row=row, value=match['gameId'])
        ws1.cell(column=2, row=row, value=match['timestamp'])
        ws1.cell(column=3, row=row, value=get_champ_name(match['champion']))
        ws1.cell(column=4, row=row, value=match['role'])
        ws1.cell(column=5, row=row, value=match['lane'])
        ws1.cell(column=6, row=row, value='Solo')
        # ws1.cell(column=7, row=row, value=get_match_info(str(match['gameId']), summonerName))
        # ws1.cell(column=8, row=row, value=matchPlayers[0])
        # ws1.cell(column=9, row=row, value=matchPlayers[1])
        # ws1.cell(column=10, row=row, value=matchPlayers[2])
        # ws1.cell(column=11, row=row, value=matchPlayers[3])

    wb.save(filename=dest_filename)

if __name__ == "__main__":
    main()
