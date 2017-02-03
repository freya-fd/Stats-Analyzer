import os
from cassiopeia import riotapi
from cassiopeia.type.core.common import LoadPolicy
from cassiopeia.type.core.common import Queue
from cassiopeia.type.core.common import Season
from cassiopeia.type.api.exception import APIError
from cassiopeia.type.core.common import Role
from cassiopeia.type.core import common

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import cell

def auto_retry(api_call_method):
    """ A decorator to automatically retry 500s (Service Unavailable) and skip 400s (Bad Request) or 404s (Not Found). """
    def call_wrapper(*args, **kwargs):
        try:
            return api_call_method(*args, **kwargs)
        except APIError as error:
            # Try Again Once
            if error.error_code in [500]:
                try:
                    print("Got a 500, trying again...")
                    return api_call_method(*args, **kwargs)
                except APIError as another_error:
                    if another_error.error_code in [500, 400, 404]:
                        pass
                    else:
                        raise another_error

            # Skip
            elif error.error_code in [400, 404]:
                print("Got a 400 or 404")
                pass

            # Fatal
            else:
                raise error
    return call_wrapper

riotapi.get_match = auto_retry(riotapi.get_match)

def main():
    riotapi.set_load_policy(LoadPolicy.lazy)
    riotapi.set_region("NA")
    riotapi.print_calls(True)
    key = os.environ["DEV_KEY"]
    riotapi.set_api_key(key)
    #RANKED_FLEX_SR: flex
    #TEAM_BUILDER_RANKED_SOLO: solo
    flex = Queue.flex
    solo = Queue.ranked_solo_queue
    pres7 = Season.preseason_7
    summonerName = input('Enter the summoner you would like to get data for: ')
    print('Looking up data for', summonerName)

    summoner = riotapi.get_summoner_by_name(summonerName)
    wb = Workbook()
    dest_filename = summonerName + '_role_per_game.xlsx'

    ws1 = wb.active
    ws1.title = "Solo"

    row = 2
    ws1.cell(column=1, row=1, value="MatchID")
    ws1.cell(column=2, row=1, value="Champion")
    ws1.cell(column=3, row=1, value="Role")
    ws1.cell(column=4, row=1, value="Lane")
    ws1.cell(column=5, row=1, value="W/L")

    for match_reference in reversed(summoner.get_match_list(ranked_queues=solo, seasons=pres7)):
        match = riotapi.get_match(match_reference)
        if match is None:
            continue
        ws1.cell(column=1, row=row, value=match_reference.id)
        ws1.cell(column=2, row=row, value=match_reference.champion.key)
        ws1.cell(column=3, row=row, value=match_reference.role.value)
        ws1.cell(column=4, row=row, value=match_reference.lane.value)
        ws1.cell(column=5, row=row, value=match.participants[summoner].stats.win)
        row = row+1

    ws2 = wb.create_sheet("Flex")

    row = 2
    ws2.cell(column=1, row=1, value="MatchID")
    ws2.cell(column=2, row=1, value="Champion")
    ws2.cell(column=3, row=1, value="Role")
    ws2.cell(column=4, row=1, value="Lane")
    ws2.cell(column=5, row=1, value="W/L")

    for match_reference in reversed(summoner.match_list(ranked_queues=flex, seasons=pres7)):
        match = riotapi.get_match(match_reference)
        if match is None:
            continue
        ws2.cell(column=1, row=row, value=match_reference.id)
        ws2.cell(column=2, row=row, value=match_reference.champion.key)
        ws2.cell(column=3, row=row, value=match_reference.role.value)
        ws2.cell(column=4, row=row, value=match_reference.lane.value)
        ws2.cell(column=5, row=row, value=match.participants[summoner].stats.win)
        row = row+1

    wb.save(filename=dest_filename)

if __name__ == "__main__":
    main()
